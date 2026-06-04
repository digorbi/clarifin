"""Generate pre-baked Excel state fixtures for sequential eval scenarios.

Run after any change to parser code or fixture data:
    python skill/evals/fixtures/generate_states.py

States produced:
    excel_states/fresh/          — clean template, no transaction data
    excel_states/after_march_nb/ — March 2026 Neobank data + parser stored in hidden sheet
"""

from __future__ import annotations

import hashlib
import importlib.util
import shutil
import sys
from datetime import datetime
from pathlib import Path
from types import ModuleType

import openpyxl
from openpyxl.styles import Alignment, Font

HERE = Path(__file__).parent
ROOT = HERE.parent.parent.parent  # skill/evals/fixtures -> skill/evals -> skill -> repo root
TEMPLATE = ROOT / "skill" / "template.xlsx"
STATEMENTS = HERE / "statements"
STATES = HERE / "excel_states"

# Add repo root to sys.path so we can import scripts/
sys.path.insert(0, str(ROOT))


def _tx_hash(date: str, description: str, amount: object) -> str:
    key = f"{date}|{description}|{amount}"
    return hashlib.sha256(key.encode()).hexdigest()[:16]


def _load_module(path: Path) -> ModuleType:
    spec = importlib.util.spec_from_file_location("_parser", path)
    assert spec and spec.loader
    mod = importlib.util.module_from_spec(spec)
    sys.modules["_parser"] = mod  # must register before exec so dataclasses resolve __module__
    spec.loader.exec_module(mod)  # type: ignore[union-attr]
    return mod


def _mock_category(description: str) -> str:
    desc = description.lower()
    if any(w in desc for w in ["transfer", "from"]):
        return "Transfer"
    if any(w in desc for w in ["cafe", "restaurant", "coffee", "bar"]):
        return "Dining"
    if any(w in desc for w in ["supermarkt", "rewe", "aldi", "lidl", "market"]):
        return "Groceries"
    if any(w in desc for w in ["bvg", "train", "taxi", "uber", "ticket"]):
        return "Transport"
    if any(w in desc for w in ["netflix", "spotify", "cinema", "games"]):
        return "Entertainment"
    if any(w in desc for w in ["pharmacy", "apotheke", "doctor", "gym"]):
        return "Health"
    return "Other"


def _month_label(month: str) -> str:
    return datetime.strptime(month, "%Y-%m").strftime("%B %Y")


def _write_monthly_tab(wb: openpyxl.Workbook, month: str, transactions: list) -> None:
    ws = wb.create_sheet(month)

    ws.merge_cells("A1:D1")
    ws["A1"].value = _month_label(month)
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    for col, h in enumerate(["Date", "Description", "Amount", "Category"], 1):
        ws.cell(row=3, column=col, value=h).font = Font(bold=True)

    prev_date = ""
    row = 4
    for tx in transactions:
        if tx.date != prev_date:
            ws.cell(row=row, column=1, value=tx.date).font = Font(bold=True)
            prev_date = tx.date
        ws.cell(row=row, column=2, value=tx.description)
        amt = ws.cell(row=row, column=3, value=float(tx.amount))
        amt.font = Font(color="00AA00" if tx.amount > 0 else "000000")
        ws.cell(row=row, column=4, value=_mock_category(tx.description))
        row += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16


def _store_parser(
    wb: openpyxl.Workbook, bank: str, fmt: str, version: int, source: str
) -> None:
    ws = wb["_parsers"]
    r = ws.max_row + 1
    for key, val in [
        ("bank", bank),
        ("format", fmt),
        ("version", str(version)),
        ("code", source),
        ("---", ""),
    ]:
        ws.cell(row=r, column=1, value=key)
        ws.cell(row=r, column=2, value=val)
        r += 1


def _store_hashes(wb: openpyxl.Workbook, transactions: list) -> None:
    ws = wb["_tx_hashes"]
    row = ws.max_row + 1
    for tx in transactions:
        ws.cell(row=row, column=1, value=_tx_hash(tx.date, tx.description, tx.amount))
        row += 1


def make_fresh(dest: Path) -> None:
    dest.mkdir(parents=True, exist_ok=True)
    shutil.copy(TEMPLATE, dest / "budget.xlsx")
    print(f"  fresh → {dest / 'budget.xlsx'}")


def make_after_march_nb(dest: Path) -> None:
    dest.mkdir(parents=True, exist_ok=True)

    parser_path = ROOT / "scripts" / "extract_neobank_joint_pdf.py"
    pdf_path = STATEMENTS / "joint_nb_0326.pdf"

    mod = _load_module(parser_path)
    stmt = mod.parse_pdf(str(pdf_path))

    wb = openpyxl.load_workbook(str(TEMPLATE))
    _write_monthly_tab(wb, "2026-03", stmt.transactions)
    _store_parser(
        wb,
        bank=stmt.bank,
        fmt="Joint Account EUR Statement",
        version=1,
        source=parser_path.read_text(),
    )
    _store_hashes(wb, stmt.transactions)

    out = dest / "budget.xlsx"
    wb.save(str(out))
    print(f"  after_march_nb → {out} ({len(stmt.transactions)} tx, parser stored)")


def main() -> None:
    print("Generating Excel state fixtures...")
    make_fresh(STATES / "fresh")
    make_after_march_nb(STATES / "after_march_nb")
    print("Done.")


if __name__ == "__main__":
    main()
