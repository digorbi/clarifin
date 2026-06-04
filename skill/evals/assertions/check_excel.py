"""Validate a generated Excel budget report for structure and data correctness.

Usage:
    python check_excel.py <excel_file_or_output_dir> --check <name> [options]

Checks:
    monthly-tab      --month YYYY-MM
    settings
    hidden-parser
    tx-count         --month YYYY-MM --expected N
    income-total     --month YYYY-MM --expected AMOUNT
    expense-total    --month YYYY-MM --expected AMOUNT
    categories       --month YYYY-MM
    no-duplicates    --month YYYY-MM
    parser-count     --expected N  (number of distinct parser blocks in _parsers)
"""

from __future__ import annotations

import argparse
import sys
from decimal import Decimal
from pathlib import Path

import openpyxl
from openpyxl.workbook import Workbook


def find_excel(path: Path) -> Path | None:
    if path.suffix == ".xlsx":
        return path
    candidates = sorted(path.rglob("*.xlsx"))
    return candidates[0] if candidates else None


def _tx_rows(ws, min_row: int = 2) -> list[tuple]:
    return [
        row
        for row in ws.iter_rows(min_row=min_row, values_only=True)
        if row[2] is not None and isinstance(row[2], (int, float))
    ]


def check_monthly_tab(wb: Workbook, month: str) -> tuple[bool, str]:
    if month in wb.sheetnames:
        return True, f"tab '{month}' found"
    return False, f"tab '{month}' missing; sheets: {wb.sheetnames}"


def check_settings(wb: Workbook) -> tuple[bool, str]:
    if "Settings" not in wb.sheetnames:
        return False, "Settings tab missing"
    rows = [r for r in wb["Settings"].iter_rows(min_row=2, values_only=True) if r[0]]
    if not rows:
        return False, "Settings tab has no categories"
    return True, f"{len(rows)} categories defined"


def check_hidden_parser(wb: Workbook) -> tuple[bool, str]:
    if "_parsers" in wb.sheetnames:
        ws = wb["_parsers"]
        has_code = any(
            r[0] == "code" for r in ws.iter_rows(values_only=True) if r[0]
        )
        if has_code:
            return True, "_parsers sheet with code rows found"
        return False, "_parsers sheet exists but has no 'code' rows"
    hidden = [n for n in wb.sheetnames if wb[n].sheet_state == "hidden"]
    parser_sheets = [h for h in hidden if "parser" in h.lower()]
    if parser_sheets:
        return True, f"hidden parser sheets: {parser_sheets}"
    return False, f"no _parsers or hidden parser sheet; hidden: {hidden}"


def check_tx_count(wb: Workbook, month: str, expected: int) -> tuple[bool, str]:
    if month not in wb.sheetnames:
        return False, f"tab '{month}' not found"
    rows = _tx_rows(wb[month])
    if len(rows) == expected:
        return True, f"{len(rows)} transactions"
    return False, f"expected {expected}, found {len(rows)}"


def check_income_total(wb: Workbook, month: str, expected: str) -> tuple[bool, str]:
    if month not in wb.sheetnames:
        return False, f"tab '{month}' not found"
    total = sum(Decimal(str(r[2])) for r in _tx_rows(wb[month]) if r[2] > 0)
    exp = Decimal(expected)
    if total == exp:
        return True, f"income total {total}"
    return False, f"expected {exp}, got {total}"


def check_expense_total(wb: Workbook, month: str, expected: str) -> tuple[bool, str]:
    if month not in wb.sheetnames:
        return False, f"tab '{month}' not found"
    total = sum(abs(Decimal(str(r[2]))) for r in _tx_rows(wb[month]) if r[2] < 0)
    exp = Decimal(expected)
    if total == exp:
        return True, f"expense total {total}"
    return False, f"expected {exp}, got {total}"


def check_categories(wb: Workbook, month: str) -> tuple[bool, str]:
    if month not in wb.sheetnames:
        return False, f"tab '{month}' not found"
    uncategorized = [
        r[1]
        for r in _tx_rows(wb[month])
        if not (len(r) > 3 and r[3] and str(r[3]).strip())
    ]
    if uncategorized:
        return False, f"{len(uncategorized)} uncategorized: {uncategorized[:3]}"
    return True, "all transactions categorized"


def check_no_duplicates(wb: Workbook, month: str) -> tuple[bool, str]:
    if month not in wb.sheetnames:
        return False, f"tab '{month}' not found"
    rows = _tx_rows(wb[month])
    keys = [(r[0], r[1], r[2]) for r in rows]
    dupes = len(keys) - len(set(keys))
    if dupes:
        return False, f"{dupes} duplicate rows"
    return True, f"{len(rows)} rows, no duplicates"


def check_parser_count(wb: Workbook, expected: int) -> tuple[bool, str]:
    if "_parsers" not in wb.sheetnames:
        return False, "_parsers sheet missing"
    bank_rows = [
        r for r in wb["_parsers"].iter_rows(values_only=True) if r[0] == "bank"
    ]
    count = len(bank_rows)
    if count == expected:
        return True, f"{count} parser block(s)"
    return False, f"expected {expected} parser block(s), found {count}"


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("target", help="Excel file or eval output directory")
    ap.add_argument("--check", required=True)
    ap.add_argument("--month", default="")
    ap.add_argument("--expected", default="")
    args = ap.parse_args()

    excel_path = find_excel(Path(args.target))
    if excel_path is None:
        print(f"FAIL: no .xlsx file found in {args.target}")
        sys.exit(1)

    wb = openpyxl.load_workbook(str(excel_path))

    dispatch = {
        "monthly-tab": lambda: check_monthly_tab(wb, args.month),
        "settings": lambda: check_settings(wb),
        "hidden-parser": lambda: check_hidden_parser(wb),
        "tx-count": lambda: check_tx_count(wb, args.month, int(args.expected)),
        "income-total": lambda: check_income_total(wb, args.month, args.expected),
        "expense-total": lambda: check_expense_total(wb, args.month, args.expected),
        "categories": lambda: check_categories(wb, args.month),
        "no-duplicates": lambda: check_no_duplicates(wb, args.month),
        "parser-count": lambda: check_parser_count(wb, int(args.expected)),
    }

    if args.check not in dispatch:
        print(f"FAIL: unknown check '{args.check}'. Options: {list(dispatch)}")
        sys.exit(1)

    passed, evidence = dispatch[args.check]()
    print(f"{'PASS' if passed else 'FAIL'} [{args.check}]: {evidence}")
    sys.exit(0 if passed else 1)


if __name__ == "__main__":
    main()
