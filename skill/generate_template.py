"""Generate template.xlsx — the blank starting Excel for a new budget file.

Run to regenerate after changing default categories or sheet structure:
    python skill/generate_template.py
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

OUTPUT = Path(__file__).parent / "template.xlsx"

DEFAULT_CATEGORIES = [
    ("Groceries", "4CAF50", "supermarket, grocery, lidl, aldi, rewe"),
    ("Dining", "FF9800", "restaurant, cafe, coffee, food delivery"),
    ("Transport", "2196F3", "uber, taxi, train, bvg, db, fuel"),
    ("Utilities", "9C27B0", "electricity, gas, water, internet, phone"),
    ("Health", "F44336", "pharmacy, doctor, gym, fitness"),
    ("Shopping", "795548", "amazon, clothing, electronics"),
    ("Entertainment", "E91E63", "cinema, netflix, spotify, games"),
    ("Transfer", "607D8B", "bank transfer, own account"),
    ("Other", "9E9E9E", ""),
]


def build_settings(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Settings")
    for col, h in enumerate(["Category", "Color (hex)", "Keywords"], 1):
        ws.cell(row=1, column=col, value=h).font = Font(bold=True)
    for row, (name, color, keywords) in enumerate(DEFAULT_CATEGORIES, 2):
        ws.cell(row=row, column=1, value=name)
        cell = ws.cell(row=row, column=2, value=f"#{color}")
        cell.fill = PatternFill("solid", fgColor=color)
        ws.cell(row=row, column=3, value=keywords)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 50


def build_parsers(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("_parsers")
    ws.sheet_state = "hidden"
    ws.cell(row=1, column=1, value="# bank statement parsers — managed by skill, do not edit manually")


def build_hashes(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("_tx_hashes")
    ws.sheet_state = "hidden"
    ws.cell(row=1, column=1, value="# sha256[:16] hashes of processed transactions")


def main() -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    build_settings(wb)
    build_parsers(wb)
    build_hashes(wb)
    wb.save(str(OUTPUT))
    print(f"Written: {OUTPUT}")


if __name__ == "__main__":
    main()
